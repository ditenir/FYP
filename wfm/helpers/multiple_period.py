import os

from django.conf import settings
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import PatternFill

from .common import *
from .ponctual_forecast import calculate as calculate_ponctual_forecast
import pandas

WEEKDAYS = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]


def calculate(dataframe, input_data):
    if input_data["aht_unit"] == "minutes":
        input_data["aht"] *= 60
    if input_data["max_waiting_time_unit"] == "minutes":
        input_data["max_waiting_time"] *= 60
    distribution_ratios = []
    for zone in set(dataframe["Zone"]):
        for language in set(dataframe["Language"]):
            for media_type in set(dataframe["Media_type"]):
                total = sum(dataframe[dataframe.Zone == zone][dataframe.Language == language][dataframe.Media_type == media_type]['Volume'])
                data = {}
                for time in set(dataframe["Time"]):
                    data[time.hour] = []
                    for weekday in WEEKDAYS:
                        if total:
                            data[time.hour].append(sum(dataframe[dataframe.Zone == zone][dataframe.Language == language][dataframe.Media_type == media_type][dataframe.Time == time][dataframe.Weekday == weekday]['Volume']) / total)
                if data[time.hour]:
                    distribution_ratios.append([(zone, language, media_type), pandas.DataFrame(data)])
    result = {}
    for i in distribution_ratios:
        for j in sorted(i[1]):
            for k in i[1][j]:
                data = input_data.copy()
                data.update({
                    "period_num": 1,
                    "working_hours_num": len(set(dataframe["Time"])),
                    "period_type": "hour"
                })
                data["calls_num"] *= k
                if i[0] not in result:
                    result[i[0]] = {}
                if j not in result[i[0]]:
                    result[i[0]][j] = []
                result[i[0]][j].append(int(find_optimal_raw_agent(data)))
    return result


def generate_excel(calculation):
    workbook = load_workbook(os.path.join(settings.EXCEL_TEMPLATES_DIRECTORY, "output/multiple_period.xlsx"))
    worksheet = workbook.active
    worksheet['B1'] = calculation.total_agents_number
    row = 5
    for table in calculation.agents_per_criteria.all():
        fill_export_with_table(row, table, worksheet)
        row += 5 + len(table.agents_table[list(table.agents_table.keys())[0]])
    worksheet.views.sheetView[0].selection[0].activeCell = "A1"
    path = "/tmp/multiple_period.xlsx"
    workbook.save(path)
    return path


def fill_export_with_table(row, table, worksheet):
    worksheet[f'B{row}'] = table.language
    worksheet[f'C{row}'] = table.zone
    worksheet[f'D{row}'] = table.media_type
    for column, weekday in enumerate(WEEKDAYS):
        worksheet[f'{chr(ord("C") + column)}{row + 2}'] = weekday
    for index, hour in enumerate(sorted(table.agents_table.keys())):
        worksheet[f'B{row + index + 3}'] = hour
        for counter, value in enumerate(table.agents_table[hour]):
            worksheet[f'{chr(ord("C") + counter)}{row + index + 3}'] = value
